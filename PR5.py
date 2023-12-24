import requests
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver import ChromeOptions
import re
import time
import pandas as pd
from openpyxl import load_workbook
from datetime import date
from openpyxl.styles import borders
from openpyxl.styles.borders import Border
import os

class City:
    def __init__(self, name, url):
        self.name = name
        self.url = url

    def getName(self):
        return self.name
    
    def getUrl(self):
        return self.url

count = 1
def get_url(url, num):


    optinons = ChromeOptions()
    optinons.add_argument('--headless=new')
    driver = webdriver.Chrome(options=optinons)
    #driver.get(url)
    driver.get(url=url)
    time.sleep(2)

    global count

    try:
        if not os.path.isdir("page_source"):
            os.mkdir("page_source")
        with open(f'page_source/page_source-{num}.html', 'w', encoding='UTF-8') as file:
            file.write(driver.page_source)
            count += 1
    except Exception as _ex:
        print(_ex)
    finally:
        driver.close()
        driver.quit()

def get_info(name, num):
    global count
# count_local = 1
# while count_local <= 1:
    temp_day = []
    time_day = []
    wind_day = []
    dir_day = []
    fall_day =[]

    time_day_1 = []
    time_day_2 = []
    time_day_3 = []
    time_day_4 = []
    temp_day_1 = []
    temp_day_2 = []
    temp_day_3 = []
    temp_day_4 = []
    wind_day_1 = []
    wind_day_2 = []
    wind_day_3 = []
    wind_day_4 = []
    dir_day_1 = []
    dir_day_2 = []
    dir_day_3 = []
    dir_day_4 = []
    fall_day_1 = []
    fall_day_2 = []
    fall_day_3 = []
    fall_day_4 = []

    night_time_1 = []
    night_time_2 = []
    night_time_3 = []
    night_temp_1 = []
    night_temp_2 = []
    night_temp_3 = []
    night_wind_1 = []
    night_wind_2 = []
    night_wind_3 = []
    night_dir_1 = []
    night_dir_2 = []
    night_dir_3 = []
    night_fall_1 = []
    night_fall_2 = []
    night_fall_3 = []

    day_time_1 = []
    day_time_2 = []
    day_time_3 = []
    day_temp_1 = []
    day_temp_2 = []
    day_temp_3 = []
    day_wind_1 = []
    day_wind_2 = []
    day_wind_3 = []
    day_dir_1 = []
    day_dir_2 = []
    day_dir_3 = []
    day_fall_1 = []
    day_fall_2 = []
    day_fall_3 = []


    with open(f'page_source/page_source-{num}.html', encoding='UTF-8') as file:
        src = file.read()
        #count_local += 1

    soup = BeautifulSoup(src, 'lxml')
    segodnya = soup.find(id = 'forecastTable_1_3')
    temp_all_day = segodnya.find_all('div', class_ = 't_0', limit= 72)
    time_all_day = segodnya.find('tr', class_ = 'forecastTime').find_all('td', 'underlineRow', limit=72)
    wind_all_day = segodnya.find_all('tr', attrs = {'style': 'line-height: 22px; vertical-align: top;'})[1].find_all('div', class_ = 'wv_0')
    direct_all_day = segodnya.find_all('td', class_ = (re.compile(r'grayLittlen underlineRow|grayLittlen2 underlineRow|grayLittled2 underlineRow|grayLittled underlineRow')))
    fall_all_day = segodnya.find_all('tr')[3].find_all('div', class_ = 'pr_0')
    
    time_day_new = []
    for i in time_all_day:
        time_day.append(i.text)
    start = 1
    stop = 79
    slice_obj = slice(start, stop)
    time_day_new = time_day[slice_obj]

    for i in temp_all_day:
        temp_day.append(i.text)
    
    for i in wind_all_day:
        if i.text == '':
            wind_day.append('0')
        else:
            wind_day.append(i.text)
    

    for i in direct_all_day:
        dir_day.append(i.text)

    for i in fall_all_day:
        string = i.attrs['onmouseover']
        string += "'[0 см]', '[0 мм]'"
        string = re.findall(r'[-+]?(?:\d*\.*\d+).см|[-+]?(?:\d*\.*\d+).мм', string)
        fall_day.append(string)

    check = 0
    check_true1 = True
    check_true2 = False
    check_true3 = False
    check_true4 = False

    for time, temp, wind, direct, fall in zip(time_day_new, temp_day, wind_day, dir_day, fall_day):
        num_time = int(time)
        if(check < num_time):
            if(check_true1):
                time_day_1.append(time)
                temp_day_1.append(temp)
                wind_day_1.append(wind)
                dir_day_1.append(direct)
                fall_day_1.append(fall)
            if(check_true2):
                time_day_2.append(time)
                temp_day_2.append(temp)
                wind_day_2.append(wind)
                dir_day_2.append(direct)
                fall_day_2.append(fall)
            if(check_true3):
                time_day_3.append(time)
                temp_day_3.append(temp)
                wind_day_3.append(wind)
                dir_day_3.append(direct)
                fall_day_3.append(fall)
            if(check_true4):
                time_day_4.append(time)
                temp_day_4.append(temp)
                wind_day_4.append(wind)
                dir_day_4.append(direct)
                fall_day_4.append(fall)
        else:
                if(check_true2 == False and check_true3 == False and check_true4 == False):
                    check_true2 = True
                    check_true1 = False
                    time_day_2.append(time)
                    temp_day_2.append(temp)
                    wind_day_2.append(wind)
                    dir_day_2.append(direct)
                    fall_day_2.append(fall)
                    check = num_time
                    #print('done2')
                    continue
                if(check_true3 == False and check_true4 == False and check_true1 == False):
                    check_true3 = True
                    check_true2 = False
                    time_day_3.append(time)
                    temp_day_3.append(temp)
                    wind_day_3.append(wind)
                    dir_day_3.append(direct)
                    fall_day_3.append(fall)
                    check = num_time
                    #print('done3')
                    continue
                if(check_true4 == False and check_true1 == False and check_true2 == False):
                    check_true4 = True
                    check_true3 = False
                    time_day_4.append(time)
                    temp_day_4.append(temp)
                    wind_day_4.append(wind)
                    dir_day_4.append(direct)
                    fall_day_4.append(fall)
                    check = num_time
                    #print('done4')
                    continue
                if(check_true4 == True):
                    #print('END')
                    break
        check = num_time

    # print('time: ', len(time_day))
    # print('temp:', len(temp_day))
    # print('Direct:', len(dir_day))
    # print(direct_all_day)

    for time, temp, wind, direct, fall in zip(time_day_1, temp_day_1, wind_day_1, dir_day_1, fall_day_1):
        if(int(time) >= 20):
            night_time_1.append(int(time))
            night_temp_1.append(int(temp))
            night_wind_1.append(int(wind))
            night_dir_1.append(direct)
        if(int(time) >=21):
            night_fall_1.append(fall)
    for time, temp, wind, direct, fall in zip(time_day_2, temp_day_2, wind_day_2, dir_day_2, fall_day_2):
        if(int(time) <= 8):
            night_time_1.append(int(time))
            night_temp_1.append(int(temp))
            night_wind_1.append(int(wind))
            night_dir_1.append(direct)
            night_fall_1.append(fall)
        if(int(time) >= 20):
            night_time_2.append(int(time))
            night_temp_2.append(int(temp))
            night_wind_2.append(int(wind))
            night_dir_2.append(direct)
        if(int(time) >= 21):
            night_fall_2.append(fall)
        if(int(time) >= 9 and int(time) <= 19):
            day_time_1.append(int(time))
            day_temp_1.append(int(temp))
            day_wind_1.append(int(wind))
            day_dir_1.append(direct)
        if(int(time) >= 9 and int(time) <= 20):
            day_fall_1.append(fall)
    for time, temp, wind, direct, fall in zip(time_day_3, temp_day_3, wind_day_3, dir_day_3, fall_day_3):
        if(int(time) <= 8):
            night_time_2.append(int(time))
            night_temp_2.append(int(temp))
            night_wind_2.append(int(wind))
            night_dir_2.append(direct)
            if (int(time) == 8):
                x = 0
                while x!=3:
                    night_fall_2.append(fall)
                    x += 1
        if(int(time) >= 20):
            night_time_3.append(int(time))
            night_temp_3.append(int(temp))
            night_wind_3.append(int(wind))
            night_dir_3.append(direct)
        if(int(time) >= 21):
            night_fall_3.append(fall)
        if(int(time) >= 9 and int(time) <= 19):
            day_time_2.append(int(time))
            day_temp_2.append(int(temp))
            day_wind_2.append(int(wind))
            day_dir_2.append(direct)
        if(int(time) >= 9 and int(time) <=20):
            day_fall_2.append(fall)
            if(int(time) >=11):
                x = 0
                while x!=3:
                    day_fall_2.append(fall)
                    x += 1
    for time, temp, wind, direct, fall in zip(time_day_4, temp_day_4, wind_day_4, dir_day_4, fall_day_4):
        if(int(time) <= 8):
            night_time_3.append(int(time))
            night_temp_3.append(int(temp))
            night_wind_3.append(int(wind))
            night_dir_3.append(direct)
            x = 0
            while x != 3:
                night_fall_3.append(fall)
                x += 1
            
        if(int(time) >= 9 and int(time) <= 19):
            day_time_3.append(int(time))
            day_temp_3.append(int(temp))
            day_wind_3.append(int(wind))
            day_dir_3.append(direct)
            if(int(time) >= 10 and int(time) <=20):
                x = 0
                while x != 3:
                    day_fall_3.append(fall)
                    x += 1

    min_night_1 = min(night_temp_1)
    min_night_2 = min(night_temp_2)
    min_night_3 = min(night_temp_3)
    max_day_1 = max(day_temp_1)
    max_day_2 = max(day_temp_2)
    max_day_3 = max(day_temp_3)

    result_fall_night_1 = osadki_rain(night_fall_1)
    result_fall_night_2 = osadki_rain(night_fall_2)
    result_fall_night_3 = osadki_rain(night_fall_3)
    result_fall_day_1 = osadki_rain(day_fall_1)
    result_fall_day_2 = osadki_rain(day_fall_2)
    result_fall_day_3 = osadki_rain(day_fall_3)

    max_wind_night_1 = get_wind(night_wind_1, night_dir_1)
    max_wind_night_2 = get_wind(night_wind_2, night_dir_2)
    max_wind_night_3 = get_wind(night_wind_3, night_dir_3)
    max_wind_day_1 = get_wind(day_wind_1, day_dir_1)
    max_wind_day_2 = get_wind(day_wind_2, day_dir_2)
    max_wind_day_3 = get_wind(day_wind_3, day_dir_3)

    Day_1 = {
        'Max' : max_day_1,
        'Min' : min_night_1,
        'Нv' : max_wind_night_1,
        'Дv': max_wind_day_1,
        'Ос.Н' : result_fall_night_1,
        'Ос.Д' : result_fall_day_1
    }
    Day_2 = {
        'Max' : max_day_2,
        'Min' : min_night_2,
        'Нv' : max_wind_night_2,
        'Дv': max_wind_day_2,
        'Ос.Н' : result_fall_night_2,
        'Ос.Д' : result_fall_day_2
    }
    Day_3 = {
        'Max' : max_day_3,
        'Min' : min_night_3,
        'Нv' : max_wind_night_3,
        'Дv': max_wind_day_3,
        'Ос.Н' : result_fall_night_3,
        'Ос.Д' : result_fall_day_3
    }

    # print(name)
    # print('Temp day', day_temp_1)
    # print('Temp night', night_temp_1)
    # print('Day-1', Day_1)
    # print('Day-2',Day_2)
    # print('Day-3',Day_3)

    InfoRP5 = [Day_1, Day_2, Day_3]
    return InfoRP5

def get_wind(list_wind, list_direct):

    maxWind = 0
    maxDirect = ' '
    for wind, direct in zip(list_wind, list_direct):
        if(wind > maxWind):
            maxWind = wind
            maxDirect = direct
    stringMax = str(maxDirect) + ', ' + str(maxWind)
    return str.upper(stringMax)

def osadki_rain(list):
    rain_total = []
    snow_total = []
    for rain in list:
        string_vr = str(rain)
        num = re.findall(r'[-+]?(?:\d*\.*\d*.мм)', string_vr)
        string_vr2 = str(num)
        num2 = re.findall(r'[-+]?(?:\d*\.*\d+)', string_vr2)
        rain_total.append(float(num2[0]))
    result = round(sum(rain_total), 2)
    
    for snow in list:
        string_vr = str(snow)
        num = re.findall(r'[-+]?(?:\d*\.*\d*.см)', string_vr)
        string_vr2 = str(num)
        num2 = re.findall(r'[-+]?(?:\d*\.*\d+)', string_vr2)
        snow_total.append(float(num2[0]))
    
    result1 = round(sum(snow_total), 2)
    result_total = round(result + result1, 3)
    return result_total  

def table_panda(listinfo):
    df = pd.DataFrame(listinfo, columns=["Min", "Max", "Нv", "Дv", "Ос.Н", "Ос.Д"], index=["Сутки 1", "Сутки 2", "Сутки 3"])
    return df

def multiple_dfs(df_list, sheets, file_name, spaces, list_name):
    # with pd.ExcelWriter(file_name, mode='a',if_sheet_exists="overlay", engine="openpyxl") as writer:
    #     row = 0
    #     for dataframe in df_list:
    #         row +=1
    #         dataframe.to_excel(writer,sheet_name=sheets,startrow=row , startcol=0)   
    #         row = row + len(dataframe.index) + spaces + 1
    #     writer._save()

    with pd.ExcelWriter(file_name, mode='w') as writer:
        
        row = 0
        for dataframe in df_list:
            row +=1
            dataframe.to_excel(writer,sheet_name=sheets,startrow=row , startcol=0)   
            row = row + len(dataframe.index) + spaces + 1

def int_name(file_name, sheets, list_name):
    wb = load_workbook(file_name)   
    today = date.today()
    date_3 = today.strftime("%d.%m.%y")
    ws = wb[sheets]
    row = 1
    row_start = 3
    row_finish = 5
    coll = 1
    border1 = borders.Side(style = None, color = 'FF000000', border_style = 'thin')
    thin = Border(left = border1, right = border1, bottom = border1, top = border1)
    for name in list_name:
        ws.cell(row,coll,name)
        ws.cell(row, 7, date_3)
        ws.cell(row, 6, 'Сегодня')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 6
    for name in list_name:
        for rows in ws.iter_rows(min_row=row_start, max_row= row_finish, min_col=1, max_col=7):
            for cell in rows:
                cell.border = thin
        row_start += 6
        row_finish += 6

    wb.save(file_name)    
    
def create_xlsx(dfs, list_city_name):
    multiple_dfs(dfs, 'RP5', 'InfoWeather.xlsx', 1, list_city_name)
    int_name('InfoWeather.xlsx', 'RP5', list_city_name)


def main():
    global count
    list_city_rp5 = [
        {'Num' : 1,'Name' : 'Южно-Сахалинск', 'Url' : 'https://rp5.ru/Погода_в_Южно-Сахалинске'},
        {'Num' : 2,'Name' : 'Холмск', 'Url' : 'https://rp5.ru/Погода_в_Холмске'},
        {'Num' : 3,'Name' : 'Корсаков', 'Url' : 'https://rp5.ru/Погода_в_Корсакове'},
        {'Num' : 4,'Name' : 'Александровск-Сахалинский', 'Url' : 'https://rp5.ru/Погода_в_Александровске-Сахалинском'},
        {'Num' : 5,'Name' : 'Тымовск', 'Url' : 'https://rp5.ru/Погода_в_Тымовском'},
        {'Num' : 6,'Name' : 'Поронайск', 'Url' : 'https://rp5.ru/Погода_в_Поронайске'},
        {'Num' : 7,'Name' : 'Северо-Курильск', 'Url' : 'https://rp5.ru/Погода_в_Северо-Курильске'},
        {'Num' : 8,'Name' : 'Курильск', 'Url' : 'https://rp5.ru/Погода_в_Курильске'},
        {'Num' : 9,'Name' : 'Оха', 'Url' : 'https://rp5.ru/Погода_в_Охе'},
        {'Num' : 10,'Name' : 'Ноглики', 'Url' : 'https://rp5.ru/Погода_в_Ногликах'},
        {'Num' : 11,'Name' : 'Южно-Курильск', 'Url' : 'https://rp5.ru/Погода_в_Южно-Курильске'},
        {'Num' : 12,'Name' : 'Углегорск', 'Url' : 'https://rp5.ru/Погода_в_Углегорске,_Сахалинская_область'},
        {'Num' : 13,'Name' : 'Ильинск', 'Url' : 'https://rp5.ru/Погода_в_Ильинском,_Сахалинская_область'},
        {'Num' : 14,'Name' : 'Макаровск', 'Url' : "https://rp5.ru/Погода_в_Макарове,_Россия"}
    ]

    for c in list_city_rp5:
        get_url(url = c['Url'], num =  c['Num'])
        print(f'--Page-{c['Name']} done --')

    list_city_new = []
    dfs = []
    list_city_name = []

    for c in list_city_rp5:
        list_city_name.append(c['Name'])
        info = get_info(name = c['Name'], num = c['Num'])
        dfs.append(table_panda(listinfo = info))
    return dfs, list_city_name


if __name__ == '__main__':
    main()
